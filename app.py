from flask import Flask,render_template,request,session,redirect,url_for
import random as rd
import sqlite3 as sq
import os
import smtplib
import pandas as pd
import openpyxl
import getpass
app = Flask(__name__)
app.secret_key=os.urandom(24)
app.database="my.db"
#======================================================================================================================================================
@app.route('/')
@app.route('/home')
@app.route('/index')
def index():
    return render_template('index.html')
#=======================================================================================================================
@app.route('/contact')
def contact():
    return render_template('contacts.html')
#=======================================================================================================================
@app.route('/contact2')
def contact2():
    if 'ml' in session:
        return render_template('contacts2.html')
    else:
        return redirect('/login')
#=======================================================================================================================
@app.route('/landing')
def landing():
    if 'ml' in session:
        return render_template('landing.html',mail=session['ml'])
    else:
        return redirect('/login')
#=======================================================================================================================
@app.route('/login')
def login():
    return render_template('login.html')
#=======================================================================================================================
@app.route('/logout')
def logout():
    session.pop('ml')
    return redirect(url_for('index'))
#=======================================================================================================================
#CALCULATION
#=======================================================================================================================
def extractmarks(lis,sheet_obj):
    for i in range(1, sheet_obj.max_row+1):
        cell_obj = sheet_obj.cell(row=i, column=2)
        if cell_obj.value == session['ml']:
            for j in range(6, (sheet_obj.max_column + 1),4):
                print(j,end=' , ')
                lis.append(sheet_obj.cell(row=i, column=j).value)
            break
    return lis
#=======================================================================================================================
def countmarks(lis):
    count = 1
    na, ans, wans = 0, 0, 0
    apt, math, tech = 0, 0, 0
    print("\nList",lis," list")
    for i in lis:
        if (count <= 10 and i == 3):
            apt = apt + 1
        elif (count <= 20 and i == 3):
            math = math + 1
        elif (i == 3 and count <=30):
            tech = tech + 1
        if i is 0:
            na = na + 1
        if i == 3:
            ans = ans + 1
        if i == -1:
            wans = wans + 1
        count = count + 1
    correct=[na,ans,wans]
    print("Aptitude",apt)
    print("Maths",math)
    print("Tech",tech)

    print("Not Answered", na)
    print("Answered", ans)
    print("Wrong Answered", wans)
    cate=[apt,math,tech]
    return correct,cate
#=======================================================================================================================
def op():
    path = "Book1.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    source1 = wb_obj.get_sheet_by_name('Sheet1')
    source2 = wb_obj.get_sheet_by_name('Sheet2')
    source3 = wb_obj.get_sheet_by_name('Sheet3')
    source4 = wb_obj.get_sheet_by_name('Sheet4')
    source5 = wb_obj.get_sheet_by_name('Sheet5')
    source6 = wb_obj.get_sheet_by_name('Sheet6')
    l1,l2,l3,l4,l5,l6 = [],[],[],[],[],[]
    l1 = extractmarks(l1,source1)
    correct1, cate1 = countmarks(l1)
    print(cate1)
    l2 = extractmarks(l2, source2)
    correct2, cate2 = countmarks(l2)
    print(cate2)
    l3 = extractmarks(l3, source3)
    correct3, cate3 = countmarks(l3)
    print(cate3)
    l4 = extractmarks(l4, source4)
    correct4, cate4 = countmarks(l4)

    l5 = extractmarks(l5, source5)
    correct5, cate5 = countmarks(l5)

    l6 = extractmarks(l6, source6)
    correct6, cate6 = countmarks(l6)
    aptitude = [None,None,None,None,None,None]
    maths = [None, None, None, None, None, None]
    technical = [None, None, None, None, None, None]
    na=[None, None, None, None, None, None]
    ans=[None, None, None, None, None, None]
    wans=[None, None, None, None, None, None]
    i=0
    while(i<6):
        if (len(cate1) is not 0):
            aptitude[0] = cate1[0]
            maths[0] = cate1[1]
            technical[0] = cate1[2]
        if (len(cate2) is not 0):
            aptitude[1] = cate2[0]
            maths[1] = cate2[1]
            technical[1] = cate2[2]
        if (len(cate3) is not 0):
            aptitude[2] = cate3[0]
            maths[2] = cate3[1]
            technical[2] = cate3[2]
        if (len(cate4) is not 0):
            aptitude[3] = cate4[0]
            maths[3] = cate4[1]
            technical[3] = cate4[2]
        if (len(cate5) is not 0):
            aptitude[4] = cate5[0]
            maths[4] = cate5[1]
            technical[4] = cate5[2]
        if (len(cate6) is not 0):
            aptitude[5] = cate6[0]
            maths[5] = cate6[1]
            technical[5] = cate6[2]
        i=i+1
    print("FInal Aptitude",aptitude)
    print(maths)
    print(technical)
    return  aptitude,maths,technical,correct1,correct2,correct3,correct4,correct5,correct6
#=======================================================================================================================
def mailing(count,Name,Email):
    if 'ml' in session:
        con = smtplib.SMTP('smtp.gmail.com', 587)
        con.starttls()
        pwd = "University@99"
        con.login('quizera2019cetpa@gmail.com', pwd)
        print(Name)
        print(count)
        msg = str("Hello "+Name+" Your Overall Score is  "+ str((count/60)*100)+" percent in all the Tests")
        #msg=str("Hello"+str(100))
        print(msg)
        con.sendmail('quizera2019cetpa@gmail.com', session['ml'], msg)
        print("Send Successfully")
        con.quit()
    else:
        return redirect('/login')
#=======================================================================================================================
@app.route('/visual')
def visual():
    if 'ml' in session:
        aptitude, maths, technical, correct1, correct2, correct3, correct4, correct5, correct6 = op()
        print("Test 1 :", correct1)
        print("Aptitude",aptitude)
        print("Marks : ", correct1[1] * 3-correct1[2]*1)
        print("Test 2 :", correct2)
        print("Marks : ", correct2[1] * 3 - correct2[2] * 1)
        print("Test 3 :", correct3)
        print("Marks : ", correct3[1] * 3 - correct3[2] * 1)
        print("Test 4 :", correct4)
        print("Marks : ", correct4[1] * 3 - correct4[2] * 1)
        print("Test 5 :", correct5)
        print("Marks : ", correct5[1] * 3 - correct5[2] * 1)
        print("Test 6 :", correct6)
        print("Marks : ", correct6[1] * 3 - correct6[2] * 1)

        return render_template('visual.html', aptitude=aptitude, maths=maths, technical=technical, correct1=correct1,
                           correct2=correct2, correct3=correct3, correct4=correct4, correct5=correct5,
                           correct6=correct6)
    else:
        return redirect('/login')
#=======================================================================================================================
@app.route('/register')
def register():
    return render_template('register.html')
#=======================================================================================================================
@app.route('/forget1')
def forget1():
    return render_template('forget.html')
#=======================================================================================================================
@app.route('/forget',methods=['POST'])
def forget():
        Email = request.form["Email"]
        Question = request.form["Question"]
        Solution = request.form["Solution"]
        conn=sq.connect("my.db")
        cursor=conn.cursor()
        print("Hello ")
        cursor.execute("SELECT * FROM datauser WHERE email=(?)",[Email])
        if cursor.fetchone() is None:
            msg = "No Data Found With The Same"
            return render_template('forget.html', msg=msg, typ="errorlogin")

        else:
            cursor.execute("SELECT pwd FROM datauser WHERE Email=(?) and ques=(?) and sol=(?)",
                           [Email,Question,Solution])
            em = cursor.fetchone()
            if (em is None):
                msg = "Data Mismatched!!Probably Your Are Hacker"
                return render_template('register.html', msg=msg, typ="alert")
            else:
                pski = em[0]
                con = smtplib.SMTP('smtp.gmail.com', 587)
                con.starttls()
                pwd = "University@99"
                con.login('quizera2019cetpa@gmail.com', pwd)
                msg = str("Your Account Password is " + str(pski))
                print(msg)
                con.sendmail('quizera2019cetpa@gmail.com', Email, msg)
                print("Send Successfully")
                con.quit()
                conn.commit()
                conn.close()
                msw="Password Send to Your Registered Email ID"
                return render_template('login.html',msw=msw,typ="forget")
#=======================================================================================================================
@app.route('/verify',methods=['POST'])
def verify():
    OTP = request.form["otp"]
    Email = request.form["Email"]
    conn = sq.connect("my.db")
    cursor = conn.cursor()
    cursor.execute("SELECT OTPOriginal FROM datauser WHERE email=(?)",[Email])
    em = cursor.fetchone()
    o=em[0]
    if(int(OTP)==o):
        print("True")
        session["ml"] = Email
        return redirect('/landing')
    else:
         cursor.execute("DELETE FROM datauser WHERE email=(?)", [Email])
         conn.commit()
         conn.close()
         return redirect('/register')
#=======================================================================================================================
@app.route('/sign_up',methods=['POST'])
def signup():
        import re
        Fname = request.form["Fname"]
        Lname = request.form["Lname"]
        Email = request.form["Email"]
        Contact_no = request.form["Contact_no"]
        Question = request.form["Question"]
        Solution = request.form["Solution"]
        Password = request.form["Password"]
        correctly=False
        if Fname.isalpha() is True and Lname.isalpha() is True and Contact_no.isdigit() is True and len(Contact_no)==10:
            correctly=True
            conn=sq.connect("my.db")
            cursor=conn.cursor()
            cursor.execute("SELECT email FROM datauser WHERE Email=(?)", [Email])
            em = cursor.fetchone()
            if(em is not None):
                msg="User Already Exist with this Email...Try with other Email ID"
                return render_template('register.html',msg=msg,typ="alert")
            else:
                otp = rd.randint(101101, 939199)
                con = smtplib.SMTP('smtp.gmail.com', 587)
                con.starttls()
                pwd = "University@99"
                con.login('quizera2019cetpa@gmail.com', pwd)
                msg = str("OTP IS "+ str(otp))
                print(msg)
                con.sendmail('quizera2019cetpa@gmail.com', Email, msg)
                print("Send Successfully")
                con.quit()
                cursor.execute("INSERT INTO datauser VALUES (?,?,?,?,?,?,?,?)",[Fname,Lname,Email,Contact_no,Question,Solution,Password,otp])
                conn.commit()
                conn.close()
                return render_template('verifyOTP.html', email=Email)
        else:
            return render_template('register.html', correctly=correctly, typ="alert")
#=======================================================================================================================
def generatequestion():
    path = "AWS.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj1 = wb_obj.get_sheet_by_name('Sheet3')
    sheet_obj2 = wb_obj.get_sheet_by_name('Sheet2')
    sheet_obj3 = wb_obj.get_sheet_by_name('Sheet1')
    l = {}
    c = 1
    answer1 = []
    set1 = rd.sample(range(2, 32), 10)
    set2 = rd.sample(range(2, 32), 10)
    set3 = rd.sample(range(2, 22), 10)
    for j in range(1, 4):
        if (j == 1):
            for i in set1:
                l[sheet_obj1.cell(row=i, column=1).value] = [sheet_obj1.cell(row=i, column=2).value,
                                                             sheet_obj1.cell(row=i, column=3).value,
                                                             sheet_obj1.cell(row=i, column=4).value,
                                                             sheet_obj1.cell(row=i, column=5).value]
                answer1.append(sheet_obj1.cell(row=i, column=6).value)
        if (j == 2):
            for i in set2:
                l[sheet_obj2.cell(row=i, column=1).value] = [sheet_obj2.cell(row=i, column=2).value,
                                                             sheet_obj2.cell(row=i, column=3).value,
                                                             sheet_obj2.cell(row=i, column=4).value,
                                                             sheet_obj2.cell(row=i, column=5).value]
                answer1.append(sheet_obj2.cell(row=i, column=6).value)
        if (j == 3):
            for i in set3:
                l[sheet_obj3.cell(row=i, column=1).value] = [sheet_obj3.cell(row=i, column=2).value,
                                                             sheet_obj3.cell(row=i, column=3).value,
                                                             sheet_obj3.cell(row=i, column=4).value,
                                                             sheet_obj3.cell(row=i, column=5).value]
                answer1.append(sheet_obj3.cell(row=i, column=6).value)

    return l,answer1

#=======================================================================================================================================================
questiolist1,questiolist2,questiolist3,questiolist4,questiolist5,questiolist6={},{},{},{},{},{}
answer1,answer2,answer3,answer4,answer5,answer6=[],[],[],[],[],[]
questiolist6, answer6 = generatequestion()
questiolist5, answer5 = generatequestion()
questiolist4, answer4 = generatequestion()
questiolist3, answer3 = generatequestion()
questiolist2, answer2 = generatequestion()
questiolist1, answer1 = generatequestion()
#=======================================================================================================================
@app.route('/signin',methods=['POST'])
def signin():
        Email = request.form["Email"]
        Password = request.form["password"]
        conn=sq.connect("my.db")
        cursor=conn.cursor()
        cursor.execute("SELECT * FROM datauser WHERE email=(?) AND pwd=(?)",[Email,Password])
        if cursor.fetchone():
            session["ml"]=Email
            return redirect('/landing')
        else:
            msg="No Data Found With The Same"
            return render_template('login.html',msg=msg,typ="errorlogin")
#=======================================================================================================================
@app.route('/test1')
def test1():
    if 'ml' in session:

        conn = sq.connect("my.db")
        cursor = conn.cursor()
        Email = ""
        Name = ""
        if 'ml' in session:
            Email = session['ml']
            cursor.execute("SELECT firstname FROM datauser WHERE  email=(?)",[session['ml']])
            em = cursor.fetchall()
            for i in em:
                Name = i[0]
        path = "Book1.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.get_sheet_by_name('Sheet1')
        print(sheet_obj.max_row)
        for i in range(1, sheet_obj.max_row+1):
            cell_obj = sheet_obj.cell(row=i, column=2)
            if cell_obj.value == session['ml']:
                return render_template("error.html")
        else:
            return render_template("test1.html", questiolist=questiolist1, ide="Test1", name=Name, lenanswer=30)
    else:
        return redirect('/login')
#=======================================================================================================================
@app.route('/test2')
def test2():
    if 'ml' in session:

        conn = sq.connect("my.db")
        cursor = conn.cursor()
        Email = ""
        Name = ""
        if 'ml' in session:
            Email = session['ml']
            cursor.execute("SELECT firstname FROM datauser WHERE  email=(?)",
                           [session['ml']])
            em = cursor.fetchall()
            for i in em:
                Name = i[0]
        path = "Book1.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.get_sheet_by_name('Sheet2')
        print(sheet_obj.max_row)
        for i in range(1, sheet_obj.max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=2)
            if cell_obj.value == session['ml']:
                return render_template("error.html")
        else:
            return render_template("test2.html", questiolist=questiolist2, ide="Test2", name=Name, lenanswer=30)
    else:
        return redirect('/login')

#=======================================================================================================================
@app.route('/test3')
def test3():
    if 'ml' in session:

        conn = sq.connect("my.db")
        cursor = conn.cursor()
        Email = ""
        Name = ""
        if 'ml' in session:
            Email = session['ml']
            cursor.execute("SELECT firstname FROM datauser WHERE  email=(?)",
                           [session['ml']])
            em = cursor.fetchall()
            for i in em:
                Name = i[0]
        path = "Book1.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.get_sheet_by_name('Sheet3')
        print(sheet_obj.max_row)
        for i in range(1, sheet_obj.max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=2)
            if cell_obj.value == session['ml']:
                return render_template("error.html")
        else:
            return render_template("test3.html", questiolist=questiolist3, ide="Test3", name=Name, lenanswer=30)
    else:
        return redirect('/login')
#=======================================================================================================================

@app.route('/test4')
def test4():
    if 'ml' in session:

        conn = sq.connect("my.db")
        cursor = conn.cursor()
        Email = ""
        Name = ""
        if 'ml' in session:
            Email = session['ml']
            cursor.execute("SELECT firstname FROM datauser WHERE  email=(?)",
                           [session['ml']])
            em = cursor.fetchall()
            for i in em:
                Name = i[0]
        path = "Book1.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.get_sheet_by_name('Sheet4')
        print(sheet_obj.max_row)
        for i in range(1, sheet_obj.max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=2)
            if cell_obj.value == session['ml']:
                return render_template("error.html")
        else:
            return render_template("test4.html", questiolist=questiolist4, ide="Test4", name=Name, lenanswer=30)
    else:
        return redirect('/login')
#=======================================================================================================================

@app.route('/test5')
def test5():
    if 'ml' in session:

        conn = sq.connect("my.db")
        cursor = conn.cursor()
        Email = ""
        Name = ""
        if 'ml' in session:
            Email = session['ml']
            cursor.execute("SELECT firstname FROM datauser WHERE  email=(?)",
                           [session['ml']])
            em = cursor.fetchall()
            for i in em:
                Name = i[0]
        path = "Book1.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.get_sheet_by_name('Sheet5')
        print(sheet_obj.max_row)
        for i in range(1, sheet_obj.max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=2)
            if cell_obj.value == session['ml']:
                return render_template("error.html")
        else:
            return render_template("test5.html", questiolist=questiolist5, ide="Test5", name=Name, lenanswer=30)

    else:
        return redirect('/login')
#=======================================================================================================================
@app.route('/test6')
def test6():
    if 'ml' in session:
        print("hbfvhebvihebvehve", answer6)
        conn = sq.connect("my.db")
        cursor = conn.cursor()
        Email = ""
        Name = ""
        if 'ml' in session:
            Email = session['ml']
            cursor.execute("SELECT firstname FROM datauser WHERE  email=(?)",
                           [session['ml']])
            em = cursor.fetchall()
            for i in em:
                Name = i[0]
        path = "Book1.xlsx"
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj.get_sheet_by_name('Sheet6')
        print(sheet_obj.max_row)
        for i in range(1, sheet_obj.max_row + 1):
            cell_obj = sheet_obj.cell(row=i, column=2)
            if cell_obj.value == session['ml']:
                return render_template("error.html")
        else:
            return render_template("test6.html", questiolist=questiolist6, ide="Test6", name=Name, lenanswer=30)
    else:
        return redirect('/login')
#=======================================================================================================================
@app.route('/submit',methods=['POST'])
def submit():
    if 'ml' in session:
        correct1, correct2, correct3, correct4, correct5, correct6 = 0, 0, 0, 0, 0, 0
        incorrect1, incorrect2, incorrect3, incorrect4, incorrect5, incorrect6 = 0, 0, 0, 0, 0, 0
        lenanswer = 30
        quest = []
        answered = []
        ide = request.form["ide"]
        if ide == "Test1":
            for i in questiolist1.keys():
                if not request.form.get(i):
                    answered.append("NoInputError#321")
                else:
                    answered.append(request.form[i])
            print("Answer 1", answered)
        elif ide == "Test2":
            for i in questiolist2.keys():
                if not request.form.get(i):
                    answered.append("NoInputError#321")
                else:
                    answered.append(request.form[i])
            print("Answer 2", answered)
        elif ide == "Test3":
            for i in questiolist3.keys():
                if not request.form.get(i):
                    answered.append("NoInputError#321")
                else:
                    answered.append(request.form[i])
            print("Answer 3", answered)
        elif ide == "Test4":
            for i in questiolist4.keys():
                if not request.form.get(i):
                    answered.append("NoInputError#321")
                else:
                    answered.append(request.form[i])
            print("Answer 4", answered)
        elif ide == "Test5":
            for i in questiolist5.keys():
                if not request.form.get(i):
                    answered.append("NoInputError#321")
                else:
                    answered.append(request.form[i])
            print("Answer 5", answered)
        elif ide == "Test6":
            for i in questiolist6.keys():
                if not request.form.get(i):
                    answered.append("NoInputError#321")
                else:
                    answered.append(request.form[i])
            print("Answer 6", answered)
        conn = sq.connect("my.db")
        cursor = conn.cursor()
        Email = session['ml']
        cursor.execute("SELECT firstname FROM datauser WHERE  email=(?)", [session['ml']])
        em = cursor.fetchall()
        for i in em:
            Name = i[0]
        cursor.close()
        conn.close()
        wb = openpyxl.load_workbook("Book1.xlsx")

        if ide == 'Test1':
            # ================================================
            sh = wb.get_sheet_by_name('Sheet1')
            questions = []
            for j in questiolist1.keys():
                questions.append(j)
            print(questions)
            sl = []
            sl.append(Name)
            sl.append(Email)
            print("Test 1")
            for i in range(30):
                sl.append(questions[i])
                sl.append(answer1[i])
                sl.append(answered[i])
                if answered[i] == answer1[i]:
                    correct1=correct1+1
                    sl.append(3)
                if answered[i] != answer1[i] and answered[i]!="NoInputError#321":
                    incorrect1 = incorrect1 + 1
                    sl.append(-1)
                if answered[i] == "NoInputError#321":
                    sl.append(0)
            print(sl)
            print(questions)
            sh.append(sl)
            wb.save('Book1.xlsx')
            wb.close()
            # ================================================
        elif ide == 'Test2':
            # ================================================
            sh = wb.get_sheet_by_name('Sheet2')
            questions = []
            for i in questiolist2.keys():
                questions.append(i)
            sl = []
            sl.append(Name)
            sl.append(Email)
            print("Test 2")
            for i in range(30):
                sl.append(questions[i])
                sl.append(answer2[i])
                sl.append(answered[i])
                if answered[i] == answer2[i]:
                    correct2 = correct2 + 1
                    sl.append(3)
                if answered[i] != answer2[i] and answered[i]!="NoInputError#321":
                    incorrect2 = incorrect2 + 1
                    sl.append(-1)
                if answered[i]=="NoInputError#321":
                    sl.append(0)
            print(sl)
            sh.append(sl)
            wb.save('Book1.xlsx')
            wb.close()
            # ================================================
        elif ide == 'Test3':
            # ================================================
            sh = wb.get_sheet_by_name('Sheet3')
            questions = []
            for i in questiolist3.keys():
                questions.append(i)
            sl = []
            sl.append(Name)
            sl.append(Email)
            print("Test 3")
            for i in range(30):
                sl.append(questions[i])
                sl.append(answer3[i])
                sl.append(answered[i])
                if answered[i] == answer3[i]:
                    correct3 = correct3 + 1
                    sl.append(3)
                if answered[i] != answer3[i] and answered[i]!="NoInputError#321":
                    incorrect3 = incorrect3 + 1
                    sl.append(-1)
                if answered[i] == "NoInputError#321":
                    sl.append(0)
            print(sl)
            sh.append(sl)
            wb.save('Book1.xlsx')
            wb.close()
            # ================================================
        elif ide == 'Test4':
            # ================================================
            sh = wb.get_sheet_by_name('Sheet4')
            questions = []
            for i in questiolist4.keys():
                questions.append(i)
            sl = []
            sl.append(Name)
            sl.append(Email)
            print("Test 4")
            for i in range(30):
                sl.append(questions[i])
                sl.append(answer4[i])
                sl.append(answered[i])
                if answered[i] == answer4[i]:
                    correct4 = correct4 + 1
                    sl.append(3)
                if answered[i] != answer4[i] and answered[i]!="NoInputError#321":
                    incorrect4 = incorrect4 + 1
                    sl.append(-1)
                if answered[i] == "NoInputError#321":
                    sl.append(0)
            print("4test", sl)
            sh.append(sl)
            wb.save('Book1.xlsx')
            wb.close()
            # ================================================
        elif ide == 'Test5':
            # ================================================
            sh = wb.get_sheet_by_name('Sheet5')
            questions = []
            for i in questiolist5.keys():
                questions.append(i)
            sl = []
            sl.append(Name)
            sl.append(Email)
            print("Test 5")
            for i in range(30):
                sl.append(questions[i])
                sl.append(answer5[i])
                sl.append(answered[i])
                if answered[i] == answer5[i]:
                    correct5 = correct5 + 1
                    sl.append(3)
                if answered[i] != answer5[i] and answered[i]!="NoInputError#321":
                    incorrect5 = incorrect5 + 1
                    sl.append(-1)
                if answered[i] == "NoInputError#321":
                    sl.append(0)
            print(sl)
            sh.append(sl)
            wb.save('Book1.xlsx')
            wb.close()
            # ================================================
        elif ide == 'Test6':
            # ================================================
            sh = wb.get_sheet_by_name('Sheet6')
            questions = []
            for i in questiolist6.keys():
                questions.append(i)
            sl = []
            sl.append(Name)
            sl.append(Email)
            print("Test 6")
            for i in range(30):
                sl.append(questions[i])
                sl.append(answer6[i])
                sl.append(answered[i])
                if answered[i] == answer6[i]:
                    correct6 = correct6 + 1
                    sl.append(3)
                if answered[i] != answer6[i] and answered[i]!="NoInputError#321":
                    incorrect6 = incorrect6 + 1
                    sl.append(-1)
                if answered[i] == "NoInputError#321":
                    sl.append(0)
            print(sl)
            sh.append(sl)
            wb.save('Book1.xlsx')
            wb.close()
            # ================================================
        if (ide == "Test1"):
            return render_template('result.html', q=zip(questiolist1.keys(), answer1), name=Name, answer=answer1,
                                   answered=answered, lenanswer=lenanswer, correct=(correct1*3)-incorrect1)
        elif (ide == "Test2"):
            return render_template('result.html', q=zip(questiolist2.keys(), answer2), name=Name, answer=answer2,
                                   answered=answered,
                                   lenanswer=lenanswer, correct=(correct2*3)-incorrect2)
        elif (ide == "Test3"):
            return render_template('result.html', q=zip(questiolist3.keys(), answer3), name=Name, answer=answer3,
                                   answered=answered,
                                   lenanswer=lenanswer, correct=(correct3*3)-incorrect3)
        elif (ide == "Test4"):
            return render_template('result.html', q=zip(questiolist4.keys(), answer4), name=Name, answer=answer4,
                                   answered=answered,
                                   lenanswer=lenanswer, correct=(correct4*3)-incorrect4)
        elif (ide == "Test5"):
            return render_template('result.html', q=zip(questiolist5.keys(), answer5), name=Name, answer=answer5,
                                   answered=answered,
                                   lenanswer=lenanswer, correct=(correct5*3)-incorrect5)
        elif (ide == "Test6"):
            return render_template('result.html', q=zip(questiolist6.keys(), answer6), name=Name, answer=answer6,
                                   answered=answered,
                                   lenanswer=lenanswer, correct=(correct6*3)-incorrect6)
    else:
        return redirect('/login')

# =======================================================================================================================
'''if __name__ == '__main__':
    app.run(host='0.0.0.0',port='80')
#=======================================================================================================================
'''
app.run()